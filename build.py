#!/usr/bin/env python3
"""
PUFC United Match Dashboard builder.

Usage:
    1. Put each match workbook (must contain a "Raw data" sheet) in ./matches/
       Optional: a "Dashboard (IND)" pivot sheet is used for player minutes.
    2. python3 build.py
    3. Deploy the generated index.html (plus the fonts/ folder) to Vercel.

Everything is computed from the Raw data sheet:
  - Team events (Row column): Goals, Chances, Crosses, Offsides, Set Plays,
    1/3 Entries, Blocks, Transitions, Create/Build the Attack, possession spans.
  - Player events: passes by type (+Pass Completed), duels (+Duel Succeeded /
    Kept Possession), defensive actions, positioning runs, shots, GK actions.
Success criteria mirror the analyst's dashboard (validated vs 11.7.26 file).
Possession time = union(In Possession) ∩ union(Ball in Play); this is a clean
interval calculation and can differ slightly from the Excel dashboard formula.

Team stats are computed per period ("1st", "2nd", "full") so the site can
toggle between them. Throw-in / corner / free-kick success rate is NOT
computed: the raw event log has no reliable retained-possession signal for
set pieces (a next-event heuristic was tested and does not reproduce the
analyst's manual dashboard numbers), so only counts are reported for those.
"""
import json, re, sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

MATCH_DIR = Path(__file__).parent / "matches"
TEMPLATE  = Path(__file__).parent / "template.html"
OUT       = Path(__file__).parent / "index.html"
ROSTER    = Path(__file__).parent / "roster.json"

def load_roster():
    """name -> 'FWD'/'MID'/'DEF'/'GK'. Not in raw data — maintained by hand in
    roster.json (create/edit it yourself; any player missing shows no position)."""
    if ROSTER.exists():
        return json.loads(ROSTER.read_text(encoding="utf-8"))
    return {}

TEAM_ROWS = {
    "In Possession","Out of Possession","Ball in Play","Out of Play",
    "Transition to Attack","Transition to Defense","Create the Attack",
    "Build the Attack","Finish the Attack","High Block","Mid Block","Low Block",
    "P'Boro Set Plays","Opp Set Plays","P'Boro Chance","Opp Chance",
    "P'Boro Cross","Opp Cross","P'Boro Goal","Opp Goal",
    "P'Boro Offside","Opp Offside","P'Boro Kick Off","Opp Kick Off",
    "P'Boro 1/3 Entry","Opp 1/3 Entry","P'Boro Touches in Box","Opp Touches in Box",
    "SUB(EVENT)",
}

PASS_TYPES = [("Forward","Forward Pass"),("Lateral","Lateral Pass"),
              ("Backward","Backward Pass"),("Long","Long Kick"),
              ("Diagonal","Diagonal Pass"),("In Behind","Balls in Behind"),
              ("Cross","Cross")]

def union(iv):
    iv = sorted(iv); out=[]
    for s,e in iv:
        if out and s <= out[-1][1]: out[-1][1] = max(out[-1][1], e)
        else: out.append([s,e])
    return out

def ilen(A,B):
    t=0
    for a0,a1 in A:
        for b0,b1 in B:
            lo,hi = max(a0,b0), min(a1,b1)
            if hi>lo: t += hi-lo
    return t

def spans(df, row, half=None):
    sub = df[df["Row"]==row]
    if half: sub = sub[sub["Half"]==half]
    return union([[r["Start time"], r["Start time"]+r["Duration"]] for _,r in sub.iterrows()])

def load_match_meta(path):
    """Read DATE/Opposition/Venue/Competition from the Dashboard sheet's
    per-half table — more reliable than parsing the Timeline string, whose
    format isn't consistent across exported files. Searches for the header
    cells by name rather than a fixed column, so layout shifts don't break it."""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return {}
    if "Dashboard" not in wb.sheetnames:
        return {}
    ws = wb["Dashboard"]
    for r in range(1, min(ws.max_row, 20)+1):
        headers = {}
        for c in range(1, ws.max_column+1):
            v = ws.cell(r, c).value
            if v in ("DATE","Opposition","Venue","Competition"):
                headers[v] = c
        if {"Opposition","Venue"} <= headers.keys():
            for rr in range(r+1, r+6):
                opp = ws.cell(rr, headers["Opposition"]).value
                venue = ws.cell(rr, headers["Venue"]).value
                if opp and venue:
                    date = ws.cell(rr, headers["DATE"]).value if "DATE" in headers else None
                    comp = ws.cell(rr, headers["Competition"]).value if "Competition" in headers else None
                    return {"opponent": str(opp).strip(), "venue": str(venue).strip().upper()[:1],
                            "date": date, "competition": comp}
    return {}

def clean_opponent(raw):
    """Dashboard's Opposition cell is sometimes filled with the whole match
    title (e.g. '15:16s vs Wolves 18.07.26') instead of just the team name —
    pull out just the opponent name in that case."""
    s = str(raw).strip()
    m = re.search(r"vs\.?\s+(.+?)(?:\s+\d{1,2}[\.\/]\d{1,2}[\.\/]\d{2,4})?\s*$", s, re.IGNORECASE)
    return m.group(1).strip() if m else s

def parse_team(fname):
    """U15 files are named with a 'U15_' (or 'U15-') prefix, e.g.
    'U15_260718_vs_Wolves.xlsx'. Anything without that prefix is U16 —
    this keeps every existing match file working unchanged."""
    return "U15" if re.match(r"^u15[_\-\s]", fname, re.IGNORECASE) else "U16"

def parse_meta(path, df, fname):
    team = parse_team(fname)
    # strip a leading 'U15_' before date-parsing, so the filename's YYMMDD
    # date convention still matches regardless of the team prefix
    fname_dated = re.sub(r"^u15[_\-\s]", "", fname, flags=re.IGNORECASE)
    extra = load_match_meta(path)
    # date: prefer the filename's leading YYMMDD (the analyst's consistent
    # naming convention across weeks) over Timeline text or Dashboard's DATE
    # cell, both of which have been inconsistently formatted/blank.
    date_iso, date_label = "", ""
    fm = re.match(r"(\d{2})(\d{2})(\d{2})", fname_dated)
    if fm:
        yy, mm, dd = fm.groups()
        date_iso = f"20{yy}-{mm}-{dd}"
        date_label = f"{int(dd)}.{int(mm)}.{yy}"
    else:
        tl = str(df["Timeline"].dropna().iloc[0])
        dm = re.search(r"([\d]{1,2})\.([\d]{1,2})\.([\d]{2,4})", tl)
        if dm:
            dd, mm, yy = dm.groups()
            yy = yy if len(yy)==4 else f"20{yy}"
            date_label = f"{int(dd)}.{int(mm)}.{yy[-2:]}"
            date_iso = f"{yy}-{int(mm):02d}-{int(dd):02d}"

    if extra.get("opponent") and extra.get("venue"):
        return {"opponent": clean_opponent(extra["opponent"]), "venue": extra["venue"],
                "dateLabel": date_label, "date": date_iso, "file": fname, "team": team}
    # full fallback: parse opponent/venue from the Timeline string (older
    # files without the Dashboard Opposition/Venue table)
    tl = str(df["Timeline"].dropna().iloc[0])
    m = re.search(r"vs\s+(.+?)\s*\((H|A)\)\s*([\d.]+)", tl)
    opp, venue = (m.group(1), m.group(2)) if m else (tl, "?")
    return {"opponent": opp, "venue": venue, "dateLabel": date_label, "date": date_iso, "file": fname, "team": team}

def chance_block(series):
    s = series.fillna("")
    c = lambda k: int(s.str.contains(k, regex=False).sum())
    return {"n": int(len(s)), "on": c("On Target"), "off": c("Off Target"),
            "blocked": c("Blocked"), "box": c("Shots in the box"),
            "openPlay": c("Shots from Open Play"), "transition": c("Shots from Transition"),
            "setPlay": c("Shots from Set play"), "restart": c("Shots from Restart"),
            "crossTo": c("Cross to Shots"), "counterTo": c("Counterattack to Shots"),
            "throughTo": c("Through pass to Shots")}

def grade(d, row):
    g = d[d["Row"]==row]["Team-Grade"].fillna("")
    return {"total": int(len(g)), "plus": int((g=="+").sum()), "minus": int((g=="-").sum())}

def trans_block(d, row, mirror=False):
    t = d[d["Row"]==row]["Team-Transition Details"].fillna("")
    fin, mid, own = (int(t.str.contains(k, regex=False).sum()) for k in ["Final 1/3","Middle 1/3","Own 1/3"])
    if mirror: fin, own = own, fin   # opp perspective: our final third = their own third
    counter = int(t.str.contains("Counter,", regex=False).sum() + t.str.endswith("Counter").sum())
    cpress  = int(t.str.contains("Counterpressing", regex=False).sum())
    return {"total": int(len(t)), "final": fin, "middle": mid, "own": own,
            "counter": counter, "counterpress": cpress}

def setplay_counts(d, row, pfx, mirror=False):
    s = d[d["Row"]==row]
    det = s["Team-Set Play Details"].fillna("")
    c = lambda k: int(det.str.contains(k, regex=False).sum())
    throw = s[det == f"{pfx} Throw In"]
    # zone tags (Final/Middle/Own 1/3) are always in P'Boro's fixed frame,
    # regardless of which team's row they're attached to — so an opponent
    # throw-in advancing into OUR defensive third is tagged 'Own 1/3', not
    # 'Final 1/3'. Mirror for the opponent side accordingly.
    zone = "Own 1/3" if mirror else "Final 1/3"
    throwFinal3 = int(throw["Team-Transition Details"].fillna("").str.contains(zone, regex=False).sum())
    return {"corner": c(f"{pfx} Corner"), "throwIn": c(f"{pfx} Throw In"),
            "freeKick": c(f"{pfx} Free Kick"), "goalKick": c(f"{pfx} Goal Kick"),
            "throwFinal3": throwFinal3}

def team_period(d):
    """Compute the full team stat block for a (possibly half-filtered) frame."""
    n = lambda row: int((d["Row"]==row).sum())
    gf, ga = n("P'Boro Goal"), n("Opp Goal")
    bip, oop_ = spans(d,"Ball in Play"), spans(d,"Out of Play")
    ip, op = spans(d,"In Possession"), spans(d,"Out of Possession")
    ipb, opb = ilen(ip,bip)/60, ilen(op,bip)/60
    play = sum(e-s for s,e in bip)/60 + sum(e-s for s,e in oop_)/60
    bipmin = sum(e-s for s,e in bip)/60
    possPct = round(ipb/(ipb+opb)*100,1) if ipb+opb else 0
    return {
        "goals": gf, "goalsAgainst": ga,
        "result": "W" if gf>ga else ("D" if gf==ga else "L"),
        "time": {"play": round(play,1), "bip": round(bipmin,1),
                 "bipPct": round(bipmin/play*100,1) if play else 0,
                 "inPoss": round(ipb,1)},
        "possPct": possPct,
        "chance": chance_block(d[d["Row"]=="P'Boro Chance"]["Team-Chance Details(Shots)"]),
        "oppChance": chance_block(d[d["Row"]=="Opp Chance"]["Team-Chance Details(Shots)"]),
        "cross": grade(d,"P'Boro Cross"), "oppCross": grade(d,"Opp Cross"),
        "offside": n("P'Boro Offside"), "oppOffside": n("Opp Offside"),
        "entry3": n("P'Boro 1/3 Entry"), "oppEntry3": n("Opp 1/3 Entry"),
        "touchBox": n("P'Boro Touches in Box"), "oppTouchBox": n("Opp Touches in Box"),
        "createAttack": grade(d,"Create the Attack"),
        "buildAttack": grade(d,"Build the Attack"),
        "finishAttack": grade(d,"Finish the Attack"),
        "buildVia": {k: int(d[d["Row"]=="Build the Attack"]["Team-Build the Attack Details"].fillna("").str.contains(v, regex=False).sum())
                     for k,v in [("turnover","Turnover"),("directPlay","Direct Play"),("reachOppHalf","Reach Opp Half")]},
        "blocks": {b.split()[0].lower(): grade(d,b) for b in ["High Block","Mid Block","Low Block"]},
        "pressDetail": {k: int(d["Team-Pressing Details"].fillna("").str.contains(v, regex=False).sum())
                        for k,v in [("regainOppHalf","Regain Opp Half"),("forceError","Force Error")]},
        "trans": trans_block(d,"Transition to Attack"),
        "oppTrans": trans_block(d,"Transition to Defense", mirror=True),
        "setPlay": setplay_counts(d,"P'Boro Set Plays","P'Boro"),
        "oppSetPlay": setplay_counts(d,"Opp Set Plays","Opp", mirror=True),
    }

def shots_list(df):
    """One row per recorded shot, with pitch coords, side and outcome."""
    sm = df[df["Shot Map X"].notna()].copy()
    out = []
    for _,r in sm.iterrows():
        side = "home" if r["Row"]=="P'Boro Chance" else ("away" if r["Row"]=="Opp Chance" else None)
        if side is None: continue
        det = str(r["Team-Chance Details(Shots)"] or "")
        result = "goal" if ("Goal" in det) else ("on" if "On Target" in det else ("blocked" if "Blocked" in det else "off"))
        out.append({"half": r["Half"], "side": side, "x": float(r["Shot Map X"]), "y": float(r["Shot Map Y"]), "result": result})
    return out

POSITION_MAP = {
    "GK": "GK",
    "CB": "DEF", "RB": "DEF", "LB": "DEF", "LWB": "DEF", "RWB": "DEF", "SW": "DEF",
    "CM": "MID", "CDM": "MID", "CAM": "MID", "DM": "MID", "AM": "MID", "RM": "MID", "LM": "MID",
    "LW": "FWD", "RW": "FWD", "ST": "FWD", "CF": "FWD", "SS": "FWD",
}

def load_positions_and_minutes(path):
    """Scan the 'Dashboard' sheet for its Time/Player/Duration(mins)/Position
    table (the 'Full time' rows) — not the Dashboard (IND) pivot, which only
    has player names, not real position codes. Column location isn't fixed,
    so this searches for the header cells rather than assuming a position."""
    positions, minutes = {}, {}
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return positions, minutes
    if "Dashboard" not in wb.sheetnames:
        return positions, minutes
    ws = wb["Dashboard"]
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column-2):
            if (ws.cell(r,c).value=="Time" and ws.cell(r,c+1).value=="Player"
                    and ws.cell(r,c+2).value=="Duration (mins)" and ws.cell(r,c+3).value=="Position"):
                rr = r+1
                while True:
                    half, name, dur, pos = (ws.cell(rr,c).value, ws.cell(rr,c+1).value,
                                             ws.cell(rr,c+2).value, ws.cell(rr,c+3).value)
                    if name is None:
                        break
                    if half == "Full time":
                        if dur is not None: minutes[name] = int(dur)
                        if pos: positions[name] = POSITION_MAP.get(str(pos).strip())
                    rr += 1
    return positions, minutes

def build_match(path):
    df = pd.read_excel(path, sheet_name="Raw data")
    meta = parse_meta(path, df, path.name)

    periods = {
        "full": team_period(df),
        "1st": team_period(df[df["Half"]=="1st"]),
        "2nd": team_period(df[df["Half"]=="2nd"]),
    }
    team = {"periods": periods, "shots": shots_list(df)}

    # ---- player rows ----
    prow_mask = ~df["Row"].isin(TEAM_ROWS) & df["Row"].notna()
    pdf = df[prow_mask]
    players = {}
    ipd, oopd = pdf["Player-In Possession Details"].fillna(""), pdf["Player-Out of Possession Details"].fillna("")
    pas, st   = pdf["Player-In Possession Pass Details"].fillna(""), pdf["Player-Stats"].fillna("")
    dd, sh    = pdf["Player-Out of Possession Duel Details"].fillna(""), pdf["Player-Shot Details"].fillna("")
    gk        = pdf["Player-GK Details"].fillna("")

    positions, minutes = load_positions_and_minutes(path)
    if not minutes:  # fallback: older files without the Dashboard position table
        try:
            piv = pd.read_excel(path, sheet_name="Dashboard (IND)", header=5)
            piv = piv.rename(columns={piv.columns[0]:"Player"}).set_index("Player")
            if "Sum of Duration (mins)" in piv.columns:
                minutes = {k: int(v) for k,v in piv["Sum of Duration (mins)"].dropna().items() if k!="Grand Total"}
        except Exception:
            pass

    for name in sorted(pdf["Row"].unique()):
        m = pdf["Row"]==name
        P_ipd, P_oop, P_pas, P_st, P_dd, P_sh, P_gk = (x[m] for x in (ipd,oopd,pas,st,dd,sh,gk))
        has = lambda series, k: series.str.contains(k, regex=False)
        cnt = lambda series, k: int(has(series,k).sum())
        duel = lambda key, col: [cnt(col,key), int((has(col,key)&has(P_st,"Duel Succeeded")).sum())]
        kept = lambda key: [cnt(P_ipd,key), int((has(P_ipd,key)&has(P_st,"Kept Possession")).sum())]
        passes = {}
        for lab,key in PASS_TYPES:
            mm = has(P_pas,key)
            passes[lab] = [int(mm.sum()), int((mm&has(P_st,"Pass Completed")).sum())]
        # Total Passes = Forward + Lateral + Backward only (completed + uncompleted),
        # per the analyst's explicit definition — Long/Diagonal/Cross/In-Behind/
        # Key-through passes are tracked separately and are NOT part of this total.
        fwd_t, fwd_c = passes["Forward"]; lat_t, lat_c = passes["Lateral"]; bak_t, bak_c = passes["Backward"]
        total_c = fwd_c + lat_c + bak_c
        total_u = (fwd_t-fwd_c) + (lat_t-lat_c) + (bak_t-bak_c)
        passes["Total"] = [total_c + total_u, total_c]
        # untyped deliveries (e.g. "Corner Kick", "Free Kick (Indirect)") — informational only
        typed = P_pas.apply(lambda s: s!="" and (any(t in s for _,t in PASS_TYPES) or "Key/Through pass" in s))
        passes["SetPiece"] = [int(((P_pas!="")&~typed).sum()), 0]
        p = {
            "mins": minutes.get(name),
            "passes": passes,
            "keyPass": cnt(P_pas,"Key/Through pass"),
            "received": cnt(P_st,"Pass Received"),
            "kept": cnt(P_st,"Kept Possession"),
            "lost": cnt(P_st,"Lost Possession"),
            "lostThirds": [cnt(P_st,"Lost Own 1/3"), cnt(P_st,"Lost Middle 1/3"), cnt(P_st,"Lost Final 1/3")],
            "critical": cnt(P_st,"Critical (Positive)"),
            "goals": cnt(P_st,"Goal") - cnt(P_st,"Goal Conceded"),
            "assists": cnt(P_st,"Assist"),
            "shots": [int((P_sh!="").sum()), cnt(P_sh,"On target"), cnt(P_sh,"Off target"), cnt(P_sh,"Blocked")],
            "a1v1": kept("1v1 Attack"), "offDuel": kept("Offensive duel"),
            "defDuel": duel("Ground duel",P_dd), "aerial": duel("Aerial duel",P_dd),
            "second": duel("2nd Balls",P_dd), "d1v1": duel("1v1 Defense",P_dd),
            "tackles": duel("Tackles",P_oop),
            "interception": cnt(P_oop,"Interception"), "recovery": cnt(P_oop,"Recovery Run"),
            "turnoverWon": cnt(P_oop,"Turnover Won"), "clearance": cnt(P_oop,"Clearance"),
            "fouls": cnt(P_oop,"Fouls"), "wonFouls": cnt(P_oop,"Won Fouls"),
            "manMark": cnt(P_oop,"Man-marking"),
            "accel": cnt(P_ipd,"Acceleration"), "pocket": cnt(P_ipd,"Move into Pocket"),
            "wide": cnt(P_ipd,"Wide area"), "backToGoal": cnt(P_ipd,"Back to goal"),
            "overlaps": cnt(P_ipd,"Over/Underlaps"), "runsBehind": cnt(P_ipd,"Runs in Behind"),
            "halfspace": cnt(P_ipd,"Half-space Runs"), "touchBox": cnt(P_ipd,"Touches in the box"),
        }
        g = {"save": cnt(P_gk,"Save"), "conceded": cnt(P_gk,"Conceded"),
             "onT": cnt(P_gk,"On target"), "offT": cnt(P_gk,"Off target"),
             "gk1v1": cnt(P_gk,"1v1"), "space": cnt(P_gk,"Defend the Space"),
             "line": cnt(P_gk,"Defend the line"), "counterKick": cnt(P_gk,"Counter Kick")}
        p["gk"] = g if sum(g.values())>0 else None
        players[name] = p

    roster = load_roster()
    for name, p in players.items():
        p["position"] = roster.get(name) or positions.get(name) or ("GK" if p["gk"] else None)

    return {"meta": meta, "team": team, "players": players}

def main():
    files = sorted(f for f in MATCH_DIR.glob("*.xlsx") if not f.name.startswith("~$"))
    if not files:
        sys.exit(f"No .xlsx files found in {MATCH_DIR}/")
    matches = []
    for f in files:
        print("·", f.name)
        matches.append(build_match(f))
    matches.sort(key=lambda m: m["meta"]["date"])
    data = json.dumps({"matches": matches}, ensure_ascii=False)
    html = TEMPLATE.read_text(encoding="utf-8").replace("__DATA_JSON__", data)
    OUT.write_text(html, encoding="utf-8")
    print(f"Built {OUT} ({len(matches)} match{'es' if len(matches)!=1 else ''}, {len(html)//1024} KB)")

if __name__ == "__main__":
    main()
